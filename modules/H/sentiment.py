import os
import openpyxl
import concurrent.futures

from main import main

from concurrent.futures import ThreadPoolExecutor
from ilock import ILock
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

from . import utils
from . import news

from .common import SEARCH_STATUS_COMPLETED
from .common import SEARCH_STATUS_ERROR

executor = ThreadPoolExecutor(1)


def vader_sentiment(text):
    analyzer = SentimentIntensityAnalyzer()
    return analyzer.polarity_scores(text)['compound']


def get_biphone_score_for_article(title):
    script_path = os.path.dirname(os.path.realpath(__file__))
    input_workbook_path = os.path.join(script_path, '..', 'G', 'Final document weight and count.xlsx')
    input_workbook = openpyxl.load_workbook(input_workbook_path)
    input_sheet = input_workbook['Sheet1']
    for row in input_sheet.iter_rows(min_row=1):
        if row[0].value == title:
            return row[7].value


def get_biphone_score_for_twitter_search(filename):
    title = filename.replace('.docx', '')
    return get_biphone_score_for_article(title)


def get_batch_search_keywords():
    script_path = os.path.dirname(os.path.realpath(__file__))
    input_workbook_path = os.path.join(script_path, '..', '..', 'Texts as found input', 'KeyWords.xlsx')
    keywords = []
    input_workbook = openpyxl.load_workbook(input_workbook_path)
    # We expect two sheets with keywords, first has positive words and the second negative ones
    for sheet_index in range(0, 2):
        sheet = input_workbook.worksheets[sheet_index]
        for x in range(2, sheet.max_row):
            cell = sheet.cell(row=x, column=1)
            keyword = cell.value
            if keyword is None:
                break
            keywords.append(keyword)
    return keywords


def news_biphone_scoring(search_id):
    with ILock('biphone_scoring_web'):
        try:
            print("Biphone scoring started!")

            main(no_cleaning=True)

            print("Biphone scoring completed! Will run post-scoring procedure now.")

            print('Updating documents in database')

            documents = utils.get_documents_for_search(search_id)
            for document in documents:
                doc_id = document[0];
                title = utils.slugify(document[1]);
                score = get_biphone_score_for_article(title)
                utils.update_document_biphone_score(doc_id, score)

            print('Updating the search status')
            utils.update_search_status(search_id, SEARCH_STATUS_COMPLETED)

            print('Done')
        except Exception as e:
            print(e)
            utils.update_search_status(search_id, SEARCH_STATUS_ERROR)


def run_news_biphone_scoring(search_id):
    executor.submit(news_biphone_scoring, search_id)


def twitter_biphone_scoring(search_id):
    with ILock('biphone_scoring_web'):
        try:
            print("Biphone scoring started!")

            main(no_cleaning=True)

            print("Biphone scoring completed! Will run post-scoring procedure now.")

            print('Updating twitter search results in database')

            search = utils.get_search(search_id)
            score = get_biphone_score_for_twitter_search(search[7])
            print('score is ', score)
            utils.update_search_biphone_score(search_id, score)

            print('Updating the search status')
            utils.update_search_status(search_id, SEARCH_STATUS_COMPLETED)

            print('Done')
        except Exception as e:
            print(e)
            utils.update_search_status(search_id, SEARCH_STATUS_ERROR)



def run_twitter_biphone_scoring(search_id):
    executor.submit(twitter_biphone_scoring, search_id)


def batch_news_biphone_scoring():
    with ILock('biphone_scoring_web'):
        print("Starting batch news biphone scoring!")
        print("Collecting articles...")
        try:
            keywords = get_batch_search_keywords()
            searches = []

            for keyword in keywords:
                articles_list = news.search_articles(keyword, 10)

                articles = []
                sentiment_scores = {}
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    futures = []

                    for article_summary in articles_list:
                        futures.append(executor.submit(news.get_article_with_score, article_summary['url']))

                    for future in concurrent.futures.as_completed(futures):
                        try:
                            article, score = future.result()
                            articles.append(article)
                            sentiment_scores[article.url] = score
                        except Exception as e:
                            # TODO
                            print(e)

                search_id = utils.save_news_search_to_db(keyword, articles, sentiment_scores)
                searches.append(search_id)
                for article in articles:
                    utils.save_article_file(article, sentiment_scores[article.url])

            print("Articles collected.")

            print("Starting biphone scoring!")

            main(no_cleaning=True)

            print("Biphone scoring completed! Will run post-scoring procedure now.")

            print('Updating documents in database')

            for search_id in searches:
                documents = utils.get_documents_for_search(search_id)
                for document in documents:
                    doc_id = document[0];
                    title = utils.slugify(document[1]);
                    score = get_biphone_score_for_article(title)
                    utils.update_document_biphone_score(doc_id, score)

                print('Updating the search status for search {}'.format(search_id))
                utils.update_search_status(search_id, SEARCH_STATUS_COMPLETED)

            print('Done')
        except Exception as e:
            print(e)
            utils.update_search_status(search_id, SEARCH_STATUS_ERROR)


def run_batch_news_biphone_scoring():
    executor.submit(batch_news_biphone_scoring)
