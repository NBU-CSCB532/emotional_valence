#!/usr/bin/python
import os
import sys
import argparse

import openpyxl
import pandas as pd
import dask.dataframe as dd
import numpy as np
import urllib3
from html.parser import HTMLParser
from openpyxl import load_workbook
from re import split
# !/usr/bin/python
import os
import sys
import argparse
import urllib3
from html.parser import HTMLParser
from openpyxl import load_workbook
from re import split

from database import db_utils

VERBOSE_DEBUG = False
PLAINTEXT = False
FILENAME = ""
WORD_SEPARATORS = [" ", "-", "—", "/", "[", "]"]
# The list below is unused due to issues with malformed output
CHARACTERS_TO_STRIP = "\r\n'‘’ "

# The URL which is used to look up words
baseURL = "https://www.oxfordlearnersdictionaries.com/search/english/?q="

# List of fricatives, needed to form possessive transcriptions. Source: https://pronuncian.com/introduction-to-fricatives
# Source 2 (not used): https://english.stackexchange.com/questions/5913/what-is-the-pronunciation-of-the-possessive-words-that-already-end-in-s
FRICATIVE_SOUNDS = ['v', 'f', 'ð', 'θ', 'z', 's', 'ʒ', 'ʃ', 'h']

# List of voiced consonants. Source: https://www.thoughtco.com/voiced-and-voiceless-consonants-1212092
VOICED_CONSONANTS = ['b', 'd', 'g', 'j', 'l', 'm', 'n', 'ng', 'r', 'sz', 'th', 'v', 'w', 'y', 'z']

# List of voiceless consonants. Source: https://www.thoughtco.com/voiced-and-voiceless-consonants-1212092
# Voiceless consonats which could also be found in the list of fricative sounds are discarded from the following list.
# VOICELESS_CONSONANTS = ['ch', 'f', 'k', 'p', 's', 'sh', 't', 'th']
VOICELESS_CONSONANTS = ['ch', 'f', 'k', 'p', 'sh', 't', 'th']

# List of vowels. Source: https://simple.wikipedia.org/wiki/Vowel
# Y is part of this list because the checks performed in this script only check the last letter which, if 'y', is interpreted as a vowel according to the source cited above.
VOWELS = ['a', 'а', 'e', 'i', 'o', 'u', 'y']

# A list of irregular verbs that couldn't be found at the Oxford Learner's Dictionaries.
# Source: https://www.englishpage.com/irregularverbs/irregularverbs.html#
# Also some present forms of verbs are listed below due to them being represented in a different way in the URL of OLD, e.g. sow1.
# NB! "inbreed" has been omitted
# NB! wound has two forms (both not directly found using this script) - noun and verb. The wordform has thus not been included.
# NB! Some verbs, such as "withdraw", have alternative spellings. Only one of these has been taken into consideration
IRREGULAR_VERBS = {
    "arisen": "əˈrɪzn",
    "beaten": "ˈbiːtn",
    "bid": "bɪd",
    "bidden": "ˈbɪdn",
    "bled": "bled",
    "daydreamt": "ˈdeɪdremt",
    "disproven": "ˌdɪsˈpruːvn",
    "do": "duː",
    "dreamt": "dremt",
    "dwelled": "dwelt",
    "eaten": "ˈiːtn",
    "forewent": "fɔːˈwent",
    "foregone": "fɔːˈɡɒn",
    "foresaw": "fɔːˈsɔː",
    "foreseen": "fɔːˈsiːn",
    "forgiven": "fəˈɡɪvn",
    "forsook": "forsook",
    "forsaken": "fəˈseɪkən",
    "frostbit": "ˈfrɒstbɪt",
    "frostbitten": "ˈfrɒstbɪtn",
    "hewn": "hjuːn",
    "hid": "hɪd",
    "kept": "kept",
    "leant": "lent",
    "leapt": "lept",
    "learnt": "lɜːnt",
    "lie": "laɪ",
    "lay": "leɪ",
    "mown": "məʊn",
    "partaken": "pɑːˈteɪkən",
    "risen": "ˈrɪzn",
    "sawn": "sɔːn",
    "slid": "slɪd",
    "sow": "səʊ",
    "sowed": "səʊd",
    "sown": "səʊn",
    "spelt": "spelt",
    "spilt": "spɪlt",
    "strewn": "struːn",
    "striven": "ˈstrɪvn",
    "sunburnt": "ˈsʌnbɜːnt",
    "taken": "ˈteɪkən",
    "typewrite": "taɪpraɪt",
    "typewrote": "taɪprəʊt",
    "typewritten": "taɪprɪtn",
    "waylaid": "weɪˈleɪd",
    "withdraw": "wɪðˈdrɔː",
    "withdrew": "wɪðˈdruː",
    "withdrawn": "wɪðˈdrɔːn",
    "withheld": "wɪðˈheld",
    "withstood": "wɪðˈstʊd"
}

# A table of prefixes' transcriptions obtained by manual observation of the OLD.
PREFIX_TRANSCRIPTIONS = {
    "mis": "ˌmɪsˈ",
    "over": "ˌəʊvəˈ",
    "out": "ˌaʊtˈ",
    "pre": "ˌpriːˈ",
    "re": "ˌriːˈ",
    "un": "ˌʌn",
    "under": "ˌʌndəˈ",
    "de": "diː"
}

# Since OLD is limited in terms of these, the transcriptions below have been fetched from Wiktionary. Might be incorrect.
CONTRACTIONS = {
    "i've": "aɪv",
    "you've": "juːv",
    "we've": "wiːv",
    "they've": "ðeɪv",
    "i'll": "aɪl",
    "you'll": "juːl",
    "he'll": "hɪl",
    "she'll": "ʃɪl",
    "it'll": "ˈɪtl̩",
    "we'll": "wɪl",
    "they'll": "ðeɪl",
    "i'm": "aɪm",
    "you're": "jʊə(ɹ)",
    "we're": "wɪə(ɹ)",
    "they're": "ðɛə(ɹ)",
    "i'd": "aɪd",
    "you'd": "juːd",
    "he'd": "hiːd",
    "she'd": "ʃiːd",
    "it'd": "ˈɪtəd",
    "we'd": "wiːd",
    "they'd": "ðeɪd",
    "it's": "ɪts"
}


class DictionaryParser(HTMLParser):
    def __init__(self):
        HTMLParser.__init__(self)
        # Needed to differentiate between British English and American English transcriptions:
        self.isBritish = False
        self.record = False
        self.found = dict()
        self.counter = 0;
        self.type = ""
        self.headWord = ""
        self.notFound = False
        self.recordError = False

        # Record type - remembers which of the following is being recorded:
        #   t - Type (noun / verb / adjective)
        #   w - Word
        #   h - Head word
        #   e - Error (Not Found)
        self.recordType = ''

    def handle_starttag(self, tag, attrs):
        if attrs is not None and attrs != []:
            if tag == "span":
                if attrs[0][0] == "class":
                    if attrs[0][1] == "phon":
                        if self.isBritish:
                            if self.type == "":
                                self.type = "unknown"
                                self.found[self.type] = []
                            self.record = True
                            self.recordType = 'w'
                            self.isBritish = False
                    if attrs[0][1] == "pos":
                        self.record = True
                        self.recordType = 't'

            elif tag == "div":
                for attr in attrs:
                    if attr[0] == "class":
                        if attr[1] == "phons_br":
                            self.isBritish = True
                            break
                    elif attr[0] == "id":
                        if attr[1] == "search-results":
                            self.notFound = True
                            break

            elif tag == "h1":
                for attr in attrs:
                    if attr[0] == "class":
                        if attr[1] == "headword":
                            self.record = True
                            self.recordType = 'h'
                            break

    def handle_endtag(self, tag):
        if self.record:
            self.record = False
            if self.recordType == 't':
                if self.type == "" or self.type is None:
                    self.type = "unknown"
                self.found[self.type] = []

    def handle_data(self, data):
        if self.record == True:
            if self.recordType == 't':
                self.type += data
            elif self.recordType == 'w':
                self.found[self.type].append(data.strip('/'))
            elif self.recordType == 'h':
                self.headWord = data
            elif self.recordType == 'e':
                if "No exact match found" in data:
                    self.notFound = True


def syllableCount(word):
    count = 0
    if word[0] in VOWELS:
        count += 1
    for index in range(1, len(word)):
        if word[index] in VOWELS and word[index - 1] not in VOWELS:
            count += 1
    if word.endswith("e"):
        count -= 1
    if count == 0:
        count += 1
    return count


def getKeysList(items):
    keysList = []
    for key in items.keys():
        keysList.append(key)
    return keysList


def getPluralOrThirdPerson(type, word, items):
    multipleItems = False
    lastChar = word[-1]
    lastTwoChars = word[-2:]
    pos = 0
    for item in items[type]:
        lastTranscriptionChar = items[type][pos][-1]
        lastTwoTranscriptionChar = items[type][pos][-2:]
        if (lastChar in VOICELESS_CONSONANTS or lastTwoChars in VOICELESS_CONSONANTS) or \
                (lastTranscriptionChar in VOICELESS_CONSONANTS or lastTwoTranscriptionChar in VOICELESS_CONSONANTS):
            items[type][pos] += "s"
        elif (lastChar in FRICATIVE_SOUNDS) or (lastTranscriptionChar in FRICATIVE_SOUNDS):
            items[type][pos] += "iz"
        elif (lastChar in VOWELS or lastChar in VOICED_CONSONANTS or lastTwoChars in VOICED_CONSONANTS) or \
                (
                        lastTranscriptionChar in VOWELS or lastTranscriptionChar in VOICED_CONSONANTS or lastTwoTranscriptionChar in VOICED_CONSONANTS):
            items[type][pos] += "z"
        pos += 1

    return items


def getTranscription(wordToTranscribe, wordType=None):
    data = []
    word = wordToTranscribe

    prefix_re = False
    prefix_un = False
    prefix_out = False
    prefix_mis = False
    prefix_pre = False
    prefix_over = False
    prefix_under = False
    prefix_de = False
    prefix = ""
    prefixExists = False
    prefixIterations = 1

    if word[:5] == "under":
        prefix_under = True
        prefixIterations = 2
        prefix = "under"

    elif word[:2] == "re":
        prefix_re = True
        prefixIterations = 2
        prefix = "re"

    elif word[:2] == "un":
        prefix_un = True
        prefixIterations = 2
        prefix = "un"

    elif word[:3] == "out":
        prefix_out = True
        prefixIterations = 2
        prefix = "out"

    elif word[:3] == "mis":
        prefix_mis = True
        prefixIterations = 2
        prefix = "mis"

    elif word[:3] == "pre":
        prefix_pre = True
        prefixIterations = 2
        prefix = "pre"

    elif word[:4] == "over":
        prefix_over = True
        prefixIterations = 2
        prefix = "over"

    elif word[:2] == "de":
        prefix_de = True
        prefixIterations = 2
        prefix = "de"

    for prefixIteration in range(prefixIterations):
        # If this is the second iteration, this means that the original word has a prefix and a match for the prefixed word
        # hasn't been found - therefore, remove the prefix and try adding it manually later.
        # This is needed in some cases where it might seem like there is a prefix but there actually isn't one,
        # for example "outside"
        if prefixIteration == 1:
            # word = prefix + word
            if prefix_re or prefix_un or prefix_de:
                word = word[2:]

            elif prefix_out or prefix_mis or prefix_pre:
                word = word[3:]

            elif prefix_over:
                word = word[4:]

            elif prefix_under:
                word = word[5:]

        # First, check if the word is present in the list of irregular verbs and, if so, return it.
        for irregular_verb in IRREGULAR_VERBS:
            if irregular_verb == word:
                transcription = dict()
                transcription["verb"] = []
                # Because the part of the script that takes out the results looks at the second item in the list, there should be a placeholder.
                transcription["verb"].append(None)

                if prefixIteration == 0:
                    transcription["verb"].append(IRREGULAR_VERBS[irregular_verb])
                else:
                    transcription["verb"].append(PREFIX_TRANSCRIPTIONS[prefix] + IRREGULAR_VERBS[irregular_verb])

                transcription["verb"].append(IRREGULAR_VERBS[irregular_verb])
                data.append(transcription)
                return data

        # If not, check if it's a contraction and get the actual determiner - the verb is added later.
        for contraction in CONTRACTIONS:
            if word == contraction:
                returnDict = dict()
                returnDict["determiner, contraction"] = list()
                returnDict["determiner, contraction"].append(CONTRACTIONS[contraction])
                data.append(returnDict)
                return data

        # Start querying the Oxford Learner's Dictionaries.

        # The next lines check for special conditions.
        isPossessive = False
        isPluralOrThirdPerson = False
        endsInLy = False
        endsInEr = False
        addTranscriptionLy = False
        addTranscriptionEr = False

        if word[-2:] == "'s" or word[-2:] == "’s":
            word = word[:-2]
            isPossessive = True
        elif (word[-4:] == "sses") or (word[-1] == "s" and word[-4:] != "ness"):
            isPluralOrThirdPerson = True
        if word[-2:] == "ly":
            endsInLy = True
        if word[-2:] == "er":
            endsInEr = True

        iterateURLs = True
        iteration = 1

        try:
            http = urllib3.PoolManager()

            # Normally, only a single iteration is expected. This might change if special conditions, as defined above, exist.
            while iterateURLs:
                URL = baseURL + word.strip()
                if VERBOSE_DEBUG:
                    print("---- NEW ITERATION -", iteration, "----")
                    print("[*] [DEBUG] Word:", word)
                parser = DictionaryParser()

                headers = {
                    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.2 Safari/605.1.15',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
                }

                openURL = http.request("GET", URL, headers=headers)
                if VERBOSE_DEBUG: print("[*] [DEBUG] Opening URL:", URL)
                redirectedURL = openURL.geturl()
                if VERBOSE_DEBUG: print("[*] [DEBUG] Redirecting to URL:", redirectedURL)

                if openURL.status == 200:
                    if VERBOSE_DEBUG: print("[*] [DEBUG] Got status 200")
                    parser.feed(openURL.data.decode())
                    iterateURLs = False
                    # Even though a query fails to find a word, HTTP 200 is returned. This means that either the word doesn't exist
                    # or that further local processing is required related to special conditions.
                    if parser.notFound:
                        if word[-1] == 's':
                            word = word[:-1]
                            iterateURLs = True
                        elif endsInLy:
                            word = word[:-2]
                            iterateURLs = True
                            endsInLy = False
                            addTranscriptionLy = True
                        elif endsInEr:
                            word = word[:-2]
                            iterateURLs = True
                            endsInEr = False
                            addTranscriptionEr = True

                    if bool(parser.found) and not parser.notFound:
                        items = parser.found

                        if addTranscriptionEr:
                            changeItems = getKeysList(items)
                            for i in range(len(changeItems[0])):
                                items[changeItems[0]][i] += "ə(r)"

                        if "noun" in items:
                            if iteration == 1 and isPluralOrThirdPerson and parser.headWord == word:
                                data.append(items)
                            elif isPossessive or (word[-1] == 's' and word != parser.headWord) or isPluralOrThirdPerson:
                                data.append(getPluralOrThirdPerson("noun", parser.headWord, items))
                                if VERBOSE_DEBUG: print(
                                    "[*] [DEBUG] Word is possessive or plural, getting transcription...")
                            else:
                                data.append(items)
                        elif "adjective" in items:
                            if isPossessive or (word[-1] == 's' and word != parser.headWord) or isPluralOrThirdPerson:
                                data.append(getPluralOrThirdPerson("adjective", parser.headWord, items))
                                if VERBOSE_DEBUG: print(
                                    "[*] [DEBUG] Word is possessive or plural, getting transcription...")

                            # https://youglish.com/pronounce/busiest/english offers different transcriptions for -er and -est.
                            elif parser.headWord[:-1] in word and parser.headWord != word:
                                if word[-2:] == "er":
                                    for key in items:
                                        items["adjective"][0] += "ə"
                                elif word[-3:] == "est":
                                    for item in items["adjective"]:
                                        items["adjective"][0] += "ɪst"
                                data.append(items)
                            elif addTranscriptionLy:
                                for i in range(len(items["adjective"])):
                                    items["adjective"][i] += "lɪ"
                                data.append(items)
                            else:
                                data.append(items)
                        elif "verb" in items:
                            data.append(items)
                        else:
                            if isPossessive or (word[-1] == 's' and word != parser.headWord) or isPluralOrThirdPerson:
                                data.append(getPluralOrThirdPerson(getKeysList(items)[0], parser.headWord, items))
                                if VERBOSE_DEBUG: print(
                                    "[*] [DEBUG] Word is possessive or plural, getting transcription...")
                            else:
                                data.append(items)

                    if VERBOSE_DEBUG:
                        print("[*] [DEBUG] Data so far:", data)
                        print("---- END ITERATION -", iteration, "----")
                    iteration += 1
                    parser.close()

                # If a word is not found however, there's nothing else to do.
                elif openURL.status == 404:
                    if VERBOSE_DEBUG: print("[*] [DEBUG] Got status 404")
                    if VERBOSE_DEBUG: print("[*] [DEBUG] URL", URL, "returned HTTP 404.")
                    iterateURLs = False
                # A different status code could mean anything so the script stops checkinf for this word.
                else:
                    if VERBOSE_DEBUG:
                        print("[*] [DEBUG] Got status", openURL.status, "- Stopping iterations")
                    else:
                        print("An error occurred.")
                    iterateURLs = False

        except Exception as e:
            print("Something went wrong, check the following details for more information:")
            print("Exception Type:", type(e))
            print("Arguments", e.args)
            print("Exception txt:", e)

        # If the requested word is found to have a prefix and this is the second iteration, this means that the word was earlier split
        # so the prefix transcription has to be added now.
        if prefixIteration == 1:
            for dictionary in data:
                for key in list(dictionary):
                    arrLocation = 0
                    for transcription in dictionary[key]:
                        dictionary[key][arrLocation] = PREFIX_TRANSCRIPTIONS[prefix] + transcription
                        arrLocation += 1

        if (len(data) != 0):
            if VERBOSE_DEBUG: print("[*] [DEBUG] Returning", data)
            return data
        else:
            if VERBOSE_DEBUG: print("[*] [DEBUG] Got no data to return")


# If this is a complex word, e.g. inter-change, split the words, get transcriptions for each and merge the results
def getComplexTranscription(wordsCombination):
    if "year" not in wordsCombination:
        data = getTranscription(wordsCombination)
        if (data != None):
            for dictionary in data:
                for key in dictionary:
                    for element in dictionary[key]:
                        return element

    words = split(" |-|—|/|[|]", wordsCombination)
    tempArray = ""
    for word in words:
        data = getTranscription(word)
        for dictionary in data:
            for key in dictionary:
                # If the word is a verb, find out which is its proper form and use it as output.
                if key == "verb":
                    if word[-3:] == "ing":
                        if syllableCount(word) == 1:
                            tempArray += dictionary["verb"][0]
                        else:
                            if dictionary["verb"][-1][-2:] == "ɪŋ":
                                tempArray += dictionary["verb"][-1] + " "
                            elif dictionary["verb"][-2][-2:] == "ɪŋ":
                                tempArray += dictionary["verb"][-2] + " "
                            else:
                                for transcription in dictionary["verb"]:
                                    if transcription[-2:] == "ɪŋ":
                                        tempArray += transcription + " "
                    elif word[-1] == "s":
                        if word[-2:] == "ss":
                            tempArray += dictionary["verb"][1] + " "
                        else:
                            tempArray += dictionary["verb"][2] + " "
                    elif word[-2:] == "ed":
                        tempArray += dictionary["verb"][3] + " "
                    else:
                        tempArray += dictionary["verb"][1] + " "
                # Otherwise, return all matched transcriptions.
                else:
                    tempArray += dictionary[key][0] + " "
                    break
    return tempArray.rstrip()


def updateProgress(thisWord, totalWords, word, errorCount):
    percentage = ("%.2f" % ((thisWord / totalWords) * 100))
    sys.stdout.write(
        "\rProgress: {0}/{1} ({2}%) ||| Errors: {3} ||| Current word is: {4}".format(thisWord, totalWords, percentage,
                                                                                     errorCount, word))
    sys.stdout.flush()
    sys.stdout.write("\033[K")
    sys.stdout.flush()
    # if (thisWord == totalWords):
    #     print ("\r\n")
    return


def nextColumn(currentColumn):
    return str(chr(ord(currentColumn) + 1))


# Check if the file provided via argument is a valid file.
def is_valid_file(parser, arg):
    if not os.path.exists(arg):
        parser.error("The file %s does not exist!" % arg)
        sys.exit()
    return arg


def get_word_transcriptions(raw_word):
    """Retrieves the transcriptions for the given word.
    Returns a list with the word as first element and the transcriptions as the remaining elements."""
    should_continue = True
    word = raw_word.strip(CHARACTERS_TO_STRIP).replace("’", "'").lower()
    result = [raw_word]
    try:
        for separator in WORD_SEPARATORS:
            if separator in word:
                result.append(getComplexTranscription(word))
                should_continue = False

        if not should_continue:
            return result

        transcription_dicts = getTranscription(word)
        for transcriptions in transcription_dicts:
            for wordType in transcriptions:
                transcribed = ""
                # If the word is a verb, find out which is its proper form and use it as output.
                if wordType == "verb":
                    if word[-3:] == "ing":
                        if syllableCount(word) == 1:
                            transcribed = transcriptions["verb"][0]
                        else:
                            if transcriptions["verb"][-1][-2:] == "ɪŋ":
                                transcribed = transcriptions["verb"][-1]
                            elif transcriptions["verb"][-2][-2:] == "ɪŋ":
                                transcribed = transcriptions["verb"][-2]
                            else:
                                for transcription in transcriptions["verb"]:
                                    if transcription[-2:] == "ɪŋ":
                                        transcribed = transcription

                    elif word[-1] == "s":
                        if word[-2:] == "ss":
                            transcribed = transcriptions["verb"][1]
                        else:
                            transcribed = transcriptions["verb"][2]
                    elif word[-2:] == "ed":
                        transcribed = transcriptions["verb"][3]
                    else:
                        transcribed = transcriptions["verb"][1]
                    result.append(transcribed)
                else:
                    # Otherwise, return all matched transcriptions.
                    result += [transcription for transcription in transcriptions[wordType] if transcription]
        return result
    except Exception as e:
        return [raw_word]

class RetrievalOfOxfordTranscriptions:
    def retrieve():
        argParser = argparse.ArgumentParser(
            description="Get transcriptions for words from the Oxford Learner's Dictionaries.",
            allow_abbrev=False)
        argParser.add_argument("-v", "--verbose", help="Produce additional DEBUG output.", action="store_true")
        argParser.add_argument("-t", "--plaintext",
                               help="If the file that has to be transcribed is not in Excel, use this flag.",
                               action="store_true")
        argParser.add_argument('-f', "--file",
                               help='Path to a .xls(x) or .txt file containing words to be transcribed. If .txt file, use -t flag',
                               type=lambda x: is_valid_file(argParser, x), nargs=1,
                               metavar='[File to get word forms from]')
        argParser.add_argument("-o", "--output", help="Save output to a separate file.", action="store",
                               type=argparse.FileType('w'), nargs=1)

        args = argParser.parse_args()

        VERBOSE_DEBUG = args.verbose
        PLAINTEXT = args.plaintext

        if args.file is not None:
            FILENAME = args.file[0]
        else:
            my_path = os.path.abspath(os.path.dirname(__file__))
            text_path = os.path.abspath(os.path.join(my_path, "../../", "Novel Word"))
            FILENAME = os.path.join(text_path, "Nontranscribed.xlsx")

        if FILENAME[-3:] == "txt" and not PLAINTEXT:
            print("It looks like you're pointing to a .txt file - please use -t.")
            sys.exit()
        elif FILENAME[-3:] == "xls" or FILENAME[-4:] == "xlsx" and PLAINTEXT:
            print("Looks like you're pointing to a .xls(x) file, please omit -t.")
            sys.exit()

        if PLAINTEXT:
            with open(FILENAME, 'r') as file:
                for word in file:
                    try:
                        # Ignore commented words
                        if word[0] != '#':
                            word = word.strip(CHARACTERS_TO_STRIP)
                            word = word.replace("’", "'")
                            if word != '':
                                shouldContinue = True
                                complex = False
                                for separator in WORD_SEPARATORS:
                                    if separator in word:
                                        result = getComplexTranscription(word)
                                        print(word, "(complex)")
                                        print("\t" + result)
                                        shouldContinue = False
                                if shouldContinue:
                                    word = word.lower()
                                    result = getTranscription(word)
                                    for transcriptions in result:
                                        for wordType in transcriptions:
                                            print(word + ' (' + wordType + ')')
                                            # If the word is a verb, find out which is its proper form and use it as output.
                                            if wordType == "verb":
                                                if word[-3:] == "ing":
                                                    if syllableCount(word) == 1:
                                                        print("\t" + transcriptions["verb"][0])
                                                    else:
                                                        if transcriptions["verb"][-1][-2:] == "ɪŋ":
                                                            print("\t" + transcriptions["verb"][-1])
                                                        elif transcriptions["verb"][-2][-2:] == "ɪŋ":
                                                            print("\t" + transcriptions["verb"][-2])
                                                        else:
                                                            for transcription in transcriptions["verb"]:
                                                                if transcription[-2:] == "ɪŋ":
                                                                    print("\t" + transcription)
                                                elif word[-1] == "s":
                                                    if word[-2:] == "ss":
                                                        print("\t" + transcriptions["verb"][1])
                                                    else:
                                                        print("\t" + transcriptions["verb"][2])
                                                elif word[-2:] == "ed":
                                                    print("\t" + transcriptions["verb"][3])
                                                else:
                                                    print("\t" + transcriptions["verb"][1])
                                            # Otherwise, return all matched transcriptions.
                                            else:
                                                for transcription in transcriptions[wordType]:
                                                    print("\t" + transcription)
                    except Exception as e:
                        print("An error occurred at word \"{0}\". Message: {1}".format(word, e))
        else:
            # Read the A column as pandas.Series
            series = pd.read_excel(FILENAME, header=None, usecols="A", squeeze=True)

            if series.size == 0:
                print("No novel words.")
                return

            # Partition the series using Dask
            ds = dd.from_pandas(series, npartitions=8)

            result = ds.apply(get_word_transcriptions, meta=(0, 'object'))
            output_file = os.path.abspath(
                    os.path.join(os.path.abspath(os.path.dirname(__file__)), "../../", "Novel Word", "Transcribed.xlsx"))
            output_dataframe = pd.DataFrame.from_records(result.compute())
            output_dataframe.to_excel(output_file, header=False, index=False)

        print("Novel Words’ transcriptions retrieved from Oxford Dictionary are in the file Transcribed.xlsx")

    def populate_transcribed_and_not_transcribed():
        my_path = os.path.abspath(os.path.dirname(__file__))
        transcribed_path = os.path.abspath(os.path.join(my_path, '../../', 'Novel Word', 'Transcribed.xlsx'))

        # load excel with its path
        wrkbk = openpyxl.load_workbook(transcribed_path)

        sh = wrkbk.active

        conn = db_utils.db_connect()
        cur = conn.cursor()

        insert_into_dictionary_sql = 'insert into dictionary(WordForm, TranscriptAsFound) values (?,?) '
        insert_into_stop_words_sql = """insert into stop_words(' Stop Words', comment, 'is a stop word') values (?,?,?)"""

        # iterate through excel and display data
        for row in sh.iter_rows(min_row=1):
            try:
                if row[1].value is None:
                    cur.execute(insert_into_stop_words_sql, (row[0].value, 'nonword', 'Stop word'))
                else:
                    cur.execute(insert_into_dictionary_sql, (row[0].value, row[1].value))
            except:
                pass

        conn.commit()
        my_path = os.path.abspath(os.path.dirname(__file__))

        dictionary_path = os.path.abspath(
            os.path.join(my_path, '../../', 'Dictionary', 'DICTIONARY Wordforms Transciptions.xlsx'))
        stop_words_path = os.path.abspath(os.path.join(my_path, '../../', 'Dictionary', 'Dictionary Stop Words.xlsx'))

        dictionary_query = """select * from dictionary"""
        cur.execute(dictionary_query)
        dictionary_result = cur.fetchall()
        pd.DataFrame(dictionary_result).to_excel(dictionary_path, header=['WordForm', 'TranscriptAsFound'], index=False)

        stop_words_query = """select * from stop_words"""
        cur.execute(stop_words_query)
        stop_words_result = cur.fetchall()
        pd.DataFrame(stop_words_result).to_excel(stop_words_path, header=[' Stop Words', 'comment', 'is a stop word'],
                                                 index=False)

        conn.close()

        print("Inserted all transcribed words into dictionary and all nontranscribed into stop words")
