import os
import re
import pandas as pd
from openpyxl import Workbook

from database import db_utils


class BiphoneWeights():
    def start(self):
        conn = db_utils.db_connect()
        conn.create_function('regexp', 2, lambda x, y: 1 if re.search(x, y) else 0)

        my_path = os.path.abspath(os.path.dirname(__file__))
        step_2_path = os.path.abspath(os.path.join(my_path, '..', 'F'))
        step_2 = os.path.join(step_2_path, 'Biphone Step 2.csv')

        cur = conn.cursor()

        query = """
        select
            fb.filename as Filename,
            sum(sw.r_value) as 'r_value_sum',
            (select count(1) from words w where w.'news ID' = fb.filename) as WordCount,
            (sum(sw.r_value) / (select count(1) from words w where w.'news ID' = fb.filename)) as Result,
            (sum(sw.r_value)) as Weight
        from final_biphone fb
        join syllables_weights sw on sw.Syllable_Key = fb.SyllableKey
        group by fb.filename;
        """

        cur.execute(query)
        result = cur.fetchall()

        my_path = os.path.abspath(os.path.dirname(__file__))
        path = os.path.abspath(os.path.join(my_path, 'Document weights.xlsx'))
        pathcount = os.path.abspath(os.path.join(my_path, 'SyllableCount.xlsx'))

        if len(result) > 0:
            pd.DataFrame(result).to_excel(path, header=['Filename', 'r_value SUM', 'Word Count', 'Result', 'Weight'], index=False)


        else:
            print("BiphoneWeights result is empty!")

        print("Calculated biphone weights")





        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=1).value = 'Filename'
        worksheet.cell(row=1, column=2).value = 'NonTranscribed'
        worksheet.cell(row=1, column=3).value = 'NBiphones'
        worksheet.cell(row=1, column=4).value = 'Transcribed'
        worksheet.cell(row=1, column=5).value = 'Result'

        row = 2

        df = pd.DataFrame()
        df = df.append(pd.read_csv(step_2), ignore_index=True)
        t1 = df.loc[2, 'Filename']
        t2 = pd.DataFrame()
        t2 = t2.append(pd.read_excel(path), ignore_index=True)
        ntr = 0
        trs = 0

        for r in range(len(df)):
            if r < len(df) - 1:
                if df.loc[r + 1, "Filename"] == t1:
                    if df.loc[r, "TranscriptOrStop"] == "Stop word":
                        ntr += 1
                    else:
                        fff = str(df.loc[r, 'BiphoneN'])
                        if 'b' in fff:
                            trs += 1

                else:
                    if df.loc[r, "TranscriptOrStop"] == "Stop word":
                        ntr += 1

                    else:
                        fff = str(df.loc[r, 'BiphoneN'])
                        if 'b' in fff:
                            trs += 1
                    numb = int(t2.loc[row - 2, 'Word Count'])
                    res = int(t2.loc[row - 2, 'Weight'])
                    worksheet.cell(row, 1).value = df.loc[r, 'Filename']
                    worksheet.cell(row, 2).value = ntr
                    worksheet.cell(row, 3).value = trs
                    worksheet.cell(row, 4).value = numb - ntr
                    worksheet.cell(row, 5).value = res/trs
                    row += 1
                    ntr = 0
                    trs = 0

                    t1 = df.loc[r + 1, "Filename"]

            else:
                if df.loc[r, "TranscriptOrStop"] == "Stop word":
                    ntr += 1
                else:
                    fff = str(df.loc[r, 'BiphoneN'])
                    if 'b' in fff:
                        trs += 1
                numb = int(t2.loc[row - 2, 'Word Count'])
                res = int(t2.loc[row - 2, 'Weight'])
                worksheet.cell(row, 1).value = df.loc[r, 'Filename']
                worksheet.cell(row, 2).value = ntr
                worksheet.cell(row, 3).value = trs
                worksheet.cell(row, 4).value = numb - ntr
                worksheet.cell(row, 5).value = res / trs
                row += 1
                ntr = 0
                trs = 0

        workbook.save(filename=os.path.abspath(
            os.path.join(os.path.abspath(os.path.dirname(__file__)), "SyllableCount") + '.xlsx'))

        f1 = pd.read_excel(path)
        f2 = pd.read_excel(pathcount)
        f3 = f1[['Filename', 'r_value SUM', 'Word Count']].merge(f2[['Filename', 'NonTranscribed', 'NBiphones', 'Transcribed']], on='Filename', how='left').merge(f1[['Filename', 'Weight']], on='Filename', how='left').merge(f2[['Filename', 'Result']])
        f3.to_excel(os.path.abspath(os.path.join(os.path.abspath(os.path.dirname(__file__)),'Final document weight and count.xlsx')),index=False)
