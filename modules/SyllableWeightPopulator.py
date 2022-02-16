import os
from os import path
import openpyxl
import pandas as pd
from scipy import stats
import numpy as np

from database import db_utils


class SyllableWeightPopulator:
    def find_correlation_between_vectors(self):
        # Get the script working directory
        script_path = path.dirname(path.realpath(__file__))

        # Get the path to the input xlsx file
        workbook_path = path.join(script_path, 'Zad 3A Explanation To the Linear Regressio  Module.xlsx')

        # Open the xslx file and load the sheet with the input data
        workbook = openpyxl.load_workbook(workbook_path)
        input_sheet = workbook['lines input Valence an biphones']

        valence_vector = []

        # We store the valence vector values in a list
        for cell in input_sheet[2][1:]:
            valence_vector.append(cell.value)

        syllable_keys = []
        biphone_frequencies = []

        # We store the Syllable keys and byphone frequencies in lists
        for row in input_sheet[3:input_sheet.max_row]:
            biphone_frequency_vector = []

            for cell in row[1:]:
                biphone_frequency_vector.append(cell.value)

            biphone_frequencies.append(biphone_frequency_vector)
            syllable_keys.append(row[0].value)

        # Create a new sheet for the output data
        output_sheet_name = 'Output New'
        workbook.create_sheet(output_sheet_name)
        output_sheet = workbook[output_sheet_name]

        output_sheet.cell(1, 1, 'Syllable_Key')
        output_sheet.cell(1, 2, 'r_value')
        output_sheet.cell(1, 3, 'p_value')

        # Use np array to store the valences
        x = np.nan_to_num(np.array(valence_vector, dtype=np.float64))

        # For each of the biphone frequencies we find the correlation to the valence vector values
        for i in range(len(biphone_frequencies)):
            y = np.nan_to_num(np.array(biphone_frequencies[i], dtype=np.float64))
            result = stats.linregress(x, y)

            # Write the calculated r and p values in the newly created sheet
            output_sheet.cell(i + 2, 1, syllable_keys[i])
            output_sheet.cell(i + 2, 2, result.rvalue)
            output_sheet.cell(i + 2, 3, result.pvalue)

        # We save the xlsx file
        workbook.save(workbook_path)
        print("Find correlation between two vectors finished")

    def populate(self):
        my_path = os.path.abspath(os.path.dirname(__file__))
        syllables_weights = os.path.abspath(
            os.path.join(my_path, 'Zad 3A Explanation To the Linear Regressio  Module.xlsx'))

        df = pd.read_excel(syllables_weights, sheet_name='Output New', usecols="A,B,C")
        conn = db_utils.db_connect()
        cur = conn.cursor()

        sql = 'DROP TABLE IF EXISTS syllables_weights'
        cur.execute(sql)
        conn.commit()

        df.to_sql('syllables_weights', con=conn, index=False)
        conn.commit()
        conn.close()

        print("The population of the syllable weights has finished")
