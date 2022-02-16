import os

import openpyxl
import pandas as pd

from database import db_utils
from modules.E.LazarProject.LazarScript import LazarScript


class BiphonicParser:
    def prepare_dictionary_for_biphone(self):
        # Get the script working directory
        script_path = os.path.dirname(os.path.realpath(__file__))

        # Get the path to the input xlsx file
        input_workbook_path = os.path.join(script_path, '../../', 'Novel Word', 'Transcribed.xlsx')

        # Open the xslx file and load the sheet with the input data
        input_workbook = openpyxl.load_workbook(input_workbook_path)
        input_sheet = input_workbook['Sheet1']

        # Get the path to the output xlsx file
        output_workbook_path = os.path.join(script_path, '../../', 'BiphoneInput', 'BiphoneInput.xlsx')

        # Create output workbook and worksheet
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.title = 'Wordform Phono Transcr'

        # Add column titles
        output_sheet.cell(1, 1, 'WordForm')
        output_sheet.cell(1, 2, 'Transcript To the Parser')
        output_sheet.cell(1, 3, 'TanscriptAsFound')

        # Start from the second row
        i = 3

        # Iterate over the input sheet and populate the output sheet
        for row in input_sheet.iter_rows(min_row=1):
            for cell in row[1:]:
                # Skip empty cells (words without transcriptions)
                if not (cell.value and not cell.value.isspace()):
                    continue

                output_sheet.cell(i, 1, row[0].value)
                output_sheet.cell(i, 2, _strip_transcript(cell.value))
                output_sheet.cell(i, 3, cell.value)
                i += 1

        # We save the xlsx file
        output_workbook.save(output_workbook_path)
        print("Prepared dictionary for biphone")

    def startLazarScript(self):
        lazarScript = LazarScript()
        lazarScript.start()

    def populate_word_syllables(self):
        my_path = os.path.abspath(os.path.dirname(__file__))
        output_path = os.path.abspath(os.path.join(my_path, "../../", "Dictionary of words Decomposed into biphones"))
        all_words_syllables_path = os.path.join(output_path, "Words Contain Syllables.xlsx")
        new_words_syllables_path = os.path.join(output_path, "NEW Words Contain Syllables.xls")

        words_syllables = pd.read_excel(all_words_syllables_path, sheet_name='Words Decomposition biphones',
                                        usecols="A,B,C,D,E,F")
        new_words_syllables = pd.read_excel(new_words_syllables_path, sheet_name='Result', usecols="A,B,C,D,E,F")

        conn = db_utils.db_connect()

        words_syllables.to_sql('words_syllables', con=conn, index=False, if_exists='replace')
        conn.commit()

        new_words_syllables.to_sql('words_syllables', con=conn, index=False, if_exists='append')
        conn.commit()

        cur = conn.cursor()
        sql = 'CREATE INDEX IF NOT EXISTS words_syllables_word_lower_idx ON words_syllables(lower(WordForm))'
        cur.execute(sql)

        select_distinct_words_syllables_sql = """
         SELECT * FROM words_syllables ws 
         GROUP BY ws.WordForm, ws.TranscriptSyl, ws.Syllable, ws.Sequence, ws.SyllableKey, ws.biphoneN"""
        cur.execute(select_distinct_words_syllables_sql)
        my_result = cur.fetchall()
        cur.close()

        if(len(my_result) > 0):
            pd.DataFrame(my_result).to_excel(all_words_syllables_path,
                                             header=['WordForm', 'TranscriptSyl', 'Syllable', 'Sequence', 'SyllableKey',
                                                     'biphoneN'], index=False,
                                             sheet_name='Words Decomposition biphones')
        else:
            print("BiphonicParser populate_word_syllables result is empty")

        conn.close()


def _strip_transcript(transcript):
    return transcript. \
        translate({ord(c): "" for c in "()ˌˈ’“”"}). \
        translate({ord(c): "_" for c in " "}). \
        translate({ord(c): "g" for c in "ɡ"})